from flask import Flask, jsonify, send_file
from flask_cors import CORS
import openpyxl
import os

app = Flask(__name__)
CORS(app)

# -----------------------------------------------
EXCEL_FILE = "23834 schools.xlsx"
SHEET_NAME = "Report"
# -----------------------------------------------

@app.route("/")
def home():
    return send_file("dashboard.html")

@app.route("/data")
def get_data():
    try:
        wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
        ws = wb[SHEET_NAME]
        data = []
        header_row = None
        header_col = None

        for row in ws.iter_rows(values_only=True):
            for i, cell in enumerate(row):
                if str(cell).strip() == "District":
                    header_row = list(row)
                    header_col = i
                    break
            if header_row:
                break

        if not header_row:
            return jsonify({"error": "Could not find District header in Report sheet"}), 400

        headers = [str(h).strip() if h is not None else "" for h in header_row[header_col:]]

        found_header = False
        for row in ws.iter_rows(values_only=True):
            row_slice = row[header_col:]
            if not found_header:
                if str(row_slice[0]).strip() == "District":
                    found_header = True
                continue
            if not any(row_slice):
                continue
            if str(row_slice[0]).strip().lower() in ('total', 'grand total', 'totals', '', 'none'):
                continue
            record = dict(zip(headers, row_slice))
            data.append(record)

        return jsonify(data)

    except FileNotFoundError:
        return jsonify({"error": f"File '{EXCEL_FILE}' not found."}), 404
    except KeyError:
        return jsonify({"error": f"Sheet '{SHEET_NAME}' not found."}), 404
    except Exception as e:
        return jsonify({"error": str(e)}), 500

if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(host="0.0.0.0", port=port, debug=False)
